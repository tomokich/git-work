#coding: cp932

import openpyxl
from datetime import datetime

# ���[�N�u�b�N�ƃV�[�g�̍쐬
workbook = openpyxl.Workbook()
sheet = workbook.active

sheet['B2']='������'
sheet['B4']='�������ABC'
sheet['B5']='��101-0022�@�����s���c��_�c������300'
sheet['B6']='TEL�F03-1234-5678�@FAX�F-03-1234-5678'
sheet['B7']='�S���Җ��F��؈�Y�@�l'
sheet['B10']='���i��'
sheet['F4']='No.'
sheet['F5']='���t'

sheet['G4']='0001'

# ���t�̐ݒ�
today = datetime.today()
sheet['G5']=today.strftime('%Y/%m/%d')
sheet['C9']='2'
sheet['C10']='=1+1'
sheet.cell(row=1, column=1).value = '=1+1'
sheet['D10']='�P��'
sheet['E10']='���z'

#���i�f�[�^
product = [
    ['���iA', 2, 10000],
    ['���iB', 1, 15000],
]

#���i�\��
row = 11        #���͍s
total = 0       #���v
for name, quantity, value in product:
    sheet.cell(row=row, column=2, value=name)
    sheet.cell(row=row, column=3, value=quantity)
    sheet.cell(row=row, column=4, value=value)
    sheet.cell(row=row, column=5, value=quantity*value)
    total += quantity*value
    row += 1
sheet.cell(row=row, column=5, value=total)

#���v�v�Z
row += 2
summary = [
    ['���v', total],
    ['�����', total*0.1],
    ['���v', total*1.1]
]
#���v�\��
for s, v in summary:
    sheet.cell(row=row, column=2, value=s)
    sheet.cell(row=row, column=5, value=v)
    row += 1

fn='������_' + today.strftime('%Y%m%d') + '.xlsx'
workbook.save(fn)

